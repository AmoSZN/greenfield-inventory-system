#!/usr/bin/env python3
"""
Convert Excel inventory to CSV for import
"""

import pandas as pd
import os

print("📊 Converting Excel to CSV...")

try:
    # Read the Excel file
    df = pd.read_excel('FULL INVENTORY LIST.xlsx')
    
    # Display info about the data
    print(f"✅ Found {len(df)} products")
    print(f"Columns: {', '.join(df.columns.tolist())}")
    
    # Show first few rows
    print("\nFirst 5 products:")
    print(df.head())
    
    # Save as CSV
    csv_filename = 'full_inventory_import.csv'
    df.to_csv(csv_filename, index=False)
    print(f"\n✅ Saved to {csv_filename}")
    
    # Create import-ready CSV with proper column names
    if 'Item #' in df.columns:
        # Rename columns for our system
        import_df = pd.DataFrame()
        import_df['ProductID'] = df['Item #']
        import_df['Description'] = df.get('Description', '')
        import_df['Quantity'] = df.get('Qty On Hand', 0)
        import_df['Notes'] = 'Imported from Excel'
        
        import_filename = 'ready_to_import.csv'
        import_df.to_csv(import_filename, index=False)
        print(f"✅ Created import-ready file: {import_filename}")
        
except Exception as e:
    print(f"❌ Error: {str(e)}")
    print("\nTrying alternative method...")
    
    # If pandas fails, try openpyxl directly
    try:
        from openpyxl import load_workbook
        wb = load_workbook('FULL INVENTORY LIST.xlsx')
        ws = wb.active
        
        with open('full_inventory_manual.csv', 'w', encoding='utf-8') as f:
            for row in ws.iter_rows(values_only=True):
                f.write(','.join(str(cell) if cell else '' for cell in row) + '\n')
        
        print("✅ Created CSV using openpyxl")
    except Exception as e2:
        print(f"❌ Alternative method also failed: {str(e2)}")